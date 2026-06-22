#!/usr/bin/env python3
"""一次性完成：解析Excel → 替换HTML函数体 → 更新版本号 → 验证JS语法"""
import openpyxl
import re
import subprocess
import sys
import os

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(WORK_DIR, 'index.html')
STORE_EXCEL = 'C:/Users/zhangying/Downloads/workbuddy-小程序月度业绩达成看板-截至6.21.xlsx'
GUIDE_EXCEL = 'C:/Users/zhangying/Downloads/workbuddy-导购实收金额汇总表-0621.xlsx'

def fmt_val(v):
    """格式化数值：None→null，数字保留原始，字符串加引号"""
    if v is None:
        return 'null'
    if isinstance(v, (int, float)):
        # 保留1位小数，但去掉多余的0
        if isinstance(v, float):
            return str(round(v, 2))
        return str(v)
    # 字符串
    s = str(v).strip()
    if s == '':
        return 'null'
    # 转义单引号
    s = s.replace("'", "\\'")
    return f"'{s}'"

def parse_store_excel():
    """解析门店业绩Excel，返回JS数组字符串"""
    wb = openpyxl.load_workbook(STORE_EXCEL, data_only=True)
    ws = wb.active
    
    # 找到表头行（包含"门店编号"的行）
    header_row = None
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val and '门店编号' in str(val):
                header_row = row
                break
        if header_row:
            break
    
    if not header_row:
        print("ERROR: 找不到表头行")
        sys.exit(1)
    
    print(f"Store header at row {header_row}")
    
    # 解析列索引（按实际表头文字精确匹配）
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(header_row, col).value or '').strip()
        if '门店编号' in val:
            col_map['store_no'] = col
        elif '门店名称' in val:
            col_map['store_name'] = col
        elif '正价' in val or '奥莱' in val or '类型' in val or '分类' in val:
            col_map['store_type'] = col
        elif '区经' in val:
            col_map['region'] = col
        elif '目标' in val and '达成' not in val:
            col_map['target'] = col
        elif '达成' in val and '率' not in val:
            col_map['achieved'] = col
        elif '实收' in val:
            col_map['actual_receipt'] = col
        elif '售后' in val:
            col_map['after_sales'] = col
    
    print(f"Column mapping: {col_map}")
    
    # 从表头行+2开始读数据（跳过合计行）
    stores = []
    for row in range(header_row + 1, ws.max_row + 1):
        store_no = ws.cell(row, col_map['store_no']).value
        if not store_no or str(store_no).strip() == '' or '合计' in str(store_no):
            continue
        
        store = {
            'store_no': str(store_no).strip(),
            'store_name': str(ws.cell(row, col_map['store_name']).value or '').strip(),
            'store_type': str(ws.cell(row, col_map['store_type']).value or '').strip(),
            'region': str(ws.cell(row, col_map['region']).value or '').strip(),
            'target': ws.cell(row, col_map['target']).value,
            'achieved': ws.cell(row, col_map['achieved']).value,
            'actual_receipt': ws.cell(row, col_map['actual_receipt']).value,
            'after_sales': ws.cell(row, col_map['after_sales']).value,
        }
        stores.append(store)
    
    print(f"Parsed {len(stores)} stores")
    
    # 生成JS数组
    lines = []
    for s in stores:
        line = "  { store_no:" + fmt_val(s['store_no'])
        line += ", store_name:" + fmt_val(s['store_name'])
        line += ", store_type:" + fmt_val(s['store_type'])
        line += ", region:" + fmt_val(s['region'])
        line += ", target:" + fmt_val(s['target'])
        line += ", achieved:" + fmt_val(s['achieved'])
        line += ", actual_receipt:" + fmt_val(s['actual_receipt'])
        line += ", after_sales:" + fmt_val(s['after_sales'])
        line += " },"
        lines.append(line)
    
    return lines, len(stores)

def parse_guide_excel():
    """解析导购实收Excel，返回JS数组字符串"""
    wb = openpyxl.load_workbook(GUIDE_EXCEL, data_only=True)
    ws = wb.active
    
    # 找表头行
    header_row = None
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val and '导购' in str(val):
                header_row = row
                break
        if header_row:
            break
    
    if not header_row:
        print("ERROR: 找不到导购表头行")
        sys.exit(1)
    
    print(f"Guide header at row {header_row}")
    
    # 解析列索引
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(header_row, col).value or '').strip()
        if '门店编号' in val or ('门店' in val and '编号' in val):
            col_map['storeCode'] = col
        elif '导购' in val and '名' in val:
            col_map['guideName'] = col
        elif '实收' in val or '金额' in val:
            col_map['amount'] = col
    
    # 如果没找到，用默认值
    if 'storeCode' not in col_map:
        col_map['storeCode'] = 2
    if 'guideName' not in col_map:
        col_map['guideName'] = 3
    if 'amount' not in col_map:
        col_map['amount'] = 4
    
    print(f"Column mapping: {col_map}")
    
    # 导购Excel结构：门店小计行有编号无姓名，导购行有姓名无编号（需继承编号）
    guides = []
    current_store_code = ''
    for row in range(header_row + 1, ws.max_row + 1):
        guide_name = ws.cell(row, col_map['guideName']).value
        store_code = ws.cell(row, col_map['storeCode']).value
        amount = ws.cell(row, col_map['amount']).value
        
        # 跳过总计行
        if store_code and '总计' in str(store_code):
            continue
        
        # 如果有门店编号，更新当前编号（这是门店小计行）
        if store_code and str(store_code).strip():
            current_store_code = str(store_code).strip()
            # 小计行没有导购姓名，跳过
            if not guide_name or str(guide_name).strip() == '':
                continue
        
        # 跳过没有导购姓名的行
        if not guide_name or str(guide_name).strip() == '':
            continue
        # 跳过没有金额的行
        if amount is None:
            continue
        
        guides.append({
            'storeCode': current_store_code,
            'guideName': str(guide_name).strip(),
            'amount': amount,
        })
    
    print(f"Parsed {len(guides)} guide records")
    
    # 生成JS数组
    lines = []
    for g in guides:
        line = "  { storeCode:" + fmt_val(g['storeCode'])
        line += ", guideName:" + fmt_val(g['guideName'])
        line += ", amount:" + fmt_val(g['amount'])
        line += " },"
        lines.append(line)
    
    return lines, len(guides)

def patch_html(store_lines, store_count, guide_lines, guide_count):
    """替换HTML中的两个数据函数，更新版本号"""
    with open(HTML_PATH, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 1. 替换 loadSampleData() 函数体
    # 匹配从 "function loadSampleData() {" 到下一个 "}" (函数结束)
    new_store_func = "function loadSampleData() {\n"
    new_store_func += "  storeData = [\n"
    new_store_func += '\n'.join(store_lines) + '\n'
    new_store_func += "];\n"
    new_store_func += "  saveToStorage();\n"
    new_store_func += "  refreshAll();\n"
    new_store_func += f"  showToast('✅ 截至6.21门店数据已载入（{store_count}家门店）');\n"
    new_store_func += "}"
    
    # 正则匹配整个函数（从 function 声明到闭合的 }）
    store_pattern = r'function loadSampleData\(\)\s*\{[^}]*\{[^}]*\}[^}]*\}'
    # 上面太复杂，用更可靠的方式：匹配从函数开始到 showToast 行后的 }
    store_pattern = r"function loadSampleData\(\)\s*\{[\s\S]*?showToast\([^)]*\);\s*\}"
    
    match = re.search(store_pattern, content)
    if not match:
        print("ERROR: 找不到 loadSampleData() 函数")
        sys.exit(1)
    
    print(f"loadSampleData: matched {match.end() - match.start()} chars")
    content = content[:match.start()] + new_store_func + content[match.end():]
    
    # 2. 替换 loadPresetGuideData() 函数体
    new_guide_func = "function loadPresetGuideData() {\n"
    new_guide_func += "  guideRawData = [\n"
    new_guide_func += '\n'.join(guide_lines) + '\n'
    new_guide_func += "];\n"
    new_guide_func += "  buildGuideMap();\n"
    new_guide_func += "  renderGuideSection();\n"
    new_guide_func += "  saveGuideToStorage();\n"
    new_guide_func += "}"
    
    # 匹配从函数声明到 saveGuideToStorage(); 后的 }
    guide_pattern = r"function loadPresetGuideData\(\)\s*\{[\s\S]*?saveGuideToStorage\(\);\s*\}"
    
    match2 = re.search(guide_pattern, content)
    if not match2:
        print("ERROR: 找不到 loadPresetGuideData() 函数")
        sys.exit(1)
    
    print(f"loadPresetGuideData: matched {match2.end() - match2.start()} chars")
    content = content[:match2.start()] + new_guide_func + content[match2.end():]
    
    # 3. 更新版本号和日期标识
    content = content.replace("v2.4-截至6.16", "v2.5-截至6.21")
    content = content.replace("v2.4", "v2.5")
    content = content.replace("截至6.16", "截至6.21")
    content = content.replace("截至6.11", "截至6.21")
    content = content.replace("截至6.19导购", "截至6.21导购")
    
    # 写入文件
    with open(HTML_PATH, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f"HTML patched successfully, new size: {len(content)} chars")
    return content

def validate_js(content):
    """提取<script>块并用node --check验证"""
    # 提取主script块
    match = re.search(r'<script>([\s\S]*?)</script>', content)
    if not match:
        print("ERROR: 找不到 <script> 块")
        sys.exit(1)
    
    js_code = match.group(1)
    js_path = os.path.join(WORK_DIR, '_validate.js')
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write(js_code)
    
    result = subprocess.run(
        ['node', '--check', js_path],
        capture_output=True, text=True, timeout=30
    )
    
    if result.returncode == 0:
        print("✅ JS syntax validation PASSED!")
    else:
        print("❌ JS syntax ERROR:")
        print(result.stderr)
        # 打印错误附近的行
        lines = js_code.split('\n')
        err_match = re.search(r'line (\d+)', result.stderr)
        if err_match:
            err_line = int(err_match.group(1))
            for i in range(max(0, err_line-3), min(len(lines), err_line+3)):
                print(f"  {i+1}: {lines[i][:120]}")
        sys.exit(1)
    
    # 清理临时文件
    os.remove(js_path)

def main():
    print("=" * 60)
    print("Step 1: Parse store Excel...")
    store_lines, store_count = parse_store_excel()
    
    print("\n" + "=" * 60)
    print("Step 2: Parse guide Excel...")
    guide_lines, guide_count = parse_guide_excel()
    
    print("\n" + "=" * 60)
    print("Step 3: Patch HTML...")
    content = patch_html(store_lines, store_count, guide_lines, guide_count)
    
    print("\n" + "=" * 60)
    print("Step 4: Validate JS syntax...")
    validate_js(content)
    
    print("\n" + "=" * 60)
    print(f"✅ ALL DONE! {store_count} stores + {guide_count} guides updated to v2.5-截至6.21")

if __name__ == '__main__':
    main()
